from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import time
import os
from selenium.webdriver.chrome.options import Options
# -------------------------------------------------------------------------
# -------------------------------------------------------------------------

print(f"\nWelcome to the App.\n")
time.sleep(3)
print(f"Please close the excel sheet ( price_sheet ).\n")
time.sleep(3)
print(f"Before Start, Please secure your sheet backup first.\n")
time.sleep(3)
input("Mahmoud, Press Enter to continue.\n")
time.sleep(2)
print(f"Loading...\n")
time.sleep(2)
print(f"Search may take a several minutes.......\n")

start_time = time.time()

options_list = Options()
options_list.add_argument("--headless")

ham_robot = webdriver.Chrome(options=options_list)

ham_robot.get("")
hold = WebDriverWait( ham_robot ,45)

# Check Point
ham_robot.implicitly_wait(30)


# -------------------------------------------------------------------------
# -------------------------------------------------------------------------

# Price Table.
fetched_table =  hold.until(EC.visibility_of_all_elements_located((By.XPATH,
                  "//div[@id='ArpPricingTableColumns']/div[@class='arp_allcolumnsdiv']/div")))


headers_list = []
price_list  = []

print(f"Please Wait......\n")

for each_card in fetched_table :

    card_name = each_card.find_element(By.CSS_SELECTOR,
                  "div > div > div.arpcolumnheader.has_arp_shortcode > div.arppricetablecolumntitle > div").text

    card_price = each_card.find_element(By.CSS_SELECTOR,
                "div > div > div.arpcolumnheader.has_arp_shortcode > div.arppricetablecolumnprice.default > div > span.arp_price_value").text

    # card_currency = each_card.find_element(By.CSS_SELECTOR,
    #             "div > div > div.arpcolumnheader.has_arp_shortcode > div.arppricetablecolumnprice.default > div > span.arp_price_duration").text

    time.sleep(0.5)
    headers_list.append(card_name)
    price_list.append(card_price)
    time.sleep(0.5)

# -------------------------------------------------------------------------
# -------------------------------------------------------------------------

main_title = hold.until(EC.visibility_of_element_located((By.XPATH, "//h3[contains(text(), 'text')]"))).text

def get_date(full_str) :
    net_date = []
    for char in full_str :
        if char.isdigit() or char == "/" :
            net_date.append(char)
    return "".join(net_date)


net_date = get_date(main_title)

# ---
headers_list.append("Date")
price_list.append(net_date)
time.sleep(1)
headers_list.append("Title")
price_list.append(main_title)
time.sleep(1)

# Get current local time
current_time_step_1 = time.localtime()
current_time_step_2 = time.strftime("%Y-%m-%d", current_time_step_1)

headers_list.append("Search day")
price_list.append(current_time_step_2)
# ---



print(f"Finishing.\n")
time.sleep(1.5)
print(f"Saving The Data.\n")


union_all = []
union_all.append(headers_list)
time.sleep(0.5)
union_all.append(price_list)

# Saving Data.

current_directory = os.getcwd()

sheet_path_1 = f"{current_directory}\\price_sheet.xlsx"
workbook_1 = excel.load_workbook(sheet_path_1)
worksheet_1 = workbook_1["Sheet1"]


# Append the List of lists
for per_row in union_all :
    worksheet_1.append(per_row)

workbook_1.save(sheet_path_1)


print(f"Search Done.")
print(f"Duration of the search = { int(time.time() - start_time) } Seconds.")
